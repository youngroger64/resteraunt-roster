from datetime import date, time
from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.employees.models import Department, Employee
from apps.roster.models import EmployeePattern, OpenShift, RosterWeek, Shift


class ManagerWorkbookTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("manager", password="test")
        self.client.force_login(self.user)
        self.roster = RosterWeek.objects.create(week_start=date(2026, 8, 10))
        self.fiona = Employee.objects.create(
            first_name="Fiona",
            last_name="Conway",
            department=Department.RESTAURANT,
            can_work_restaurant=True,
        )
        self.emily = Employee.objects.create(
            first_name="Emily",
            last_name="Ryan",
            department=Department.RESTAURANT,
            can_work_restaurant=True,
        )
        for employee in (self.fiona, self.emily):
            EmployeePattern.objects.create(
                employee=employee,
                normal_department=Department.RESTAURANT,
                average_weekly_hours=25,
                restaurant_target_hours=25,
                average_days_worked=4,
                day_probabilities={"mon": 100},
                typical_shifts={
                    "mon": {"shift": "08:30-17:00", "confidence": 100}
                },
            )
        Shift.objects.create(
            roster_week=self.roster,
            employee=self.fiona,
            department=Department.RESTAURANT,
            date=date(2026, 8, 11),
            start_time=time(10),
            end_time=time(16),
        )
        OpenShift.objects.create(
            roster_week=self.roster,
            department=Department.RESTAURANT,
            date=date(2026, 8, 10),
            start_time=time(8, 30),
            end_time=time(17),
            source_signature="08:30-17:00",
        )

    def test_export_merges_provisional_shift_into_employee_row(self):
        response = self.client.get(reverse("roster:excel", args=[self.roster.pk]))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Roster", "_Lists", "_Meta"])
        sheet = workbook["Roster"]
        names = [sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1)]
        self.assertEqual(names.count("Fiona Conway") + names.count("Emily Ryan"), 2)
        self.assertNotIn("Needs a choice", workbook.sheetnames)
        self.assertGreater(len(sheet.data_validations.dataValidation), 0)

    def test_import_replaces_roster_from_manager_workbook(self):
        response = self.client.get(reverse("roster:excel", args=[self.roster.pk]))
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["Roster"]
        # Find Fiona's row and change Tuesday to OFF while keeping Monday draft.
        fiona_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "Fiona Conway"
        )
        sheet.cell(fiona_row, 4).value = "OFF"
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "manager.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = self.client.post(
            reverse("roster:excel_import", args=[self.roster.pk]),
            {"workbook": upload},
        )
        self.assertEqual(result.status_code, 302)
        self.assertFalse(
            Shift.objects.filter(
                roster_week=self.roster,
                employee=self.fiona,
                date=date(2026, 8, 11),
            ).exists()
        )
        self.assertEqual(self.roster.open_shifts.count(), 0)
