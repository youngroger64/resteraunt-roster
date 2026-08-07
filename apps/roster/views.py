from datetime import timedelta
from io import BytesIO
from django.contrib import messages
import xlsxwriter
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Case, IntegerField, Value, When
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
from apps.employees.models import Department, Employee
from .forms import GeneratePatternRosterForm, RosterWeekForm
from .models import (
    EmployeePattern,
    OpenShift,
    PayrollRecord,
    PayrollWeek,
    RosterPurpose,
    RosterStatus,
    RosterWeek,
    Shift,
    StaffingPattern,
)
from .services.generator import (
    candidate_availability,
    copy_roster,
    generate_business_roster,
    parse_signature,
    rank_candidates,
    signature_duration,
)
from .services.learner import learn_patterns
from .services.publisher import publish_roster

@login_required
def roster_list(request):
    return render(request, "roster/list.html", {"rosters":RosterWeek.objects.all()})

@login_required
def roster_create(request):
    form = RosterWeekForm(request.POST or None)
    latest = RosterWeek.objects.filter(purpose=RosterPurpose.WEEKLY).first()
    if request.method == "POST" and form.is_valid():
        roster = form.save(commit=False)
        roster.purpose = RosterPurpose.WEEKLY
        roster.save()
        if request.POST.get("copy_latest") and latest:
            copy_roster(latest, roster)
        return redirect("roster:detail", pk=roster.pk)
    return render(request, "roster/create.html", {"form":form,"latest":latest})


@login_required
def learn(request):
    historic_count = RosterWeek.objects.filter(
        purpose=RosterPurpose.HISTORIC
    ).count()
    payroll_week_count = PayrollWeek.objects.count()
    payroll_record_count = PayrollRecord.objects.count()

    if request.method == "POST":
        result = learn_patterns()
        messages.success(
            request,
            (
                f"Learned {len(result['employees'])} employees, "
                f"{result['shift_templates']} recurring shift templates "
                f"and {result['staffing_patterns']} fallback staffing patterns."
            ),
        )
        return redirect("roster:patterns")

    return render(
        request,
        "roster/learn.html",
        {
            "historic_count": historic_count,
            "payroll_week_count": payroll_week_count,
            "payroll_record_count": payroll_record_count,
        },
    )
@login_required
def pattern_list(request):
    return render(request, "roster/patterns.html", {
        "patterns":EmployeePattern.objects.select_related("employee")
    })

@login_required
def generate_pattern_roster(request):
    form = GeneratePatternRosterForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        week_start = form.cleaned_data["week_start"]
        existing = RosterWeek.objects.filter(week_start=week_start).first()
        replace_existing = request.POST.get("replace_existing") == "yes"

        if existing and not replace_existing:
            return render(
                request,
                "roster/generate_patterns.html",
                {
                    "form": form,
                    "existing_roster": existing,
                },
            )

        if existing:
            if existing.status == RosterStatus.PUBLISHED:
                messages.error(
                    request,
                    "That roster is published. Choose: open it, or create another week.",
                )
                return render(
                    request,
                    "roster/generate_patterns.html",
                    {
                        "form": form,
                        "existing_roster": existing,
                        "published_existing": True,
                    },
                )

            roster = existing
            roster.shifts.all().delete()
            roster.open_shifts.all().delete()
            request.session.pop(f"unresolved_{roster.pk}", None)
        else:
            roster = RosterWeek.objects.create(
                week_start=week_start,
                purpose=RosterPurpose.WEEKLY,
            )

        threshold = (
            0
            if form.cleaned_data["uncertain_choice"] == "best"
            else 75
        )

        result = generate_business_roster(
            roster,
            uncertain_threshold=threshold,
        )

        if existing:
            messages.success(
                request,
                f"Draft replaced with {result['created']} assigned shift segments. "
                f"{result['open']} shifts need a choice.",
            )
        else:
            messages.success(
                request,
                f"Generated {result['created']} assigned shift segments. "
                f"{result['open']} shifts need a choice.",
            )

        return redirect("roster:detail", pk=roster.pk)

    return render(
        request,
        "roster/generate_patterns.html",
        {"form": form},
    )

@login_required
def roster_detail(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)
    days = [roster.week_start + timedelta(days=i) for i in range(7)]

    shifts = list(roster.shifts.select_related("employee"))
    shift_map = {}
    employee_hours = {}
    scheduled_employee_ids = set()

    for shift in shifts:
        shift_map.setdefault((shift.employee_id, shift.date), []).append(shift)
        employee_hours[shift.employee_id] = (
            employee_hours.get(shift.employee_id, 0) + shift.duration_hours
        )
        scheduled_employee_ids.add(shift.employee_id)

    show_all = request.GET.get("show") == "all"

    employees = Employee.objects.filter(is_active=True)
    if not show_all:
        employees = employees.filter(pk__in=scheduled_employee_ids)

    employees = employees.annotate(
        area_order=Case(
            When(department=Department.RESTAURANT, then=Value(0)),
            When(department=Department.BAR, then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        )
    ).order_by("area_order", "first_name", "last_name")

    patterns = list(
        EmployeePattern.objects.select_related("employee")
    )
    patterns_by_employee = {
        pattern.employee_id: pattern
        for pattern in patterns
    }

    current_hours = {}
    current_days = {}
    employee_days = set()
    for shift in shifts:
        key = (shift.employee_id, shift.department)
        current_hours[key] = current_hours.get(key, 0.0) + shift.duration_hours
        employee_days.add((shift.employee_id, shift.department, shift.date))

    for employee_id, department, shift_date in employee_days:
        key = (employee_id, department)
        current_days[key] = current_days.get(key, 0) + 1

    open_choice_groups = {
        Department.RESTAURANT: [],
        Department.BAR: [],
    }

    for open_shift in roster.open_shifts.all():
        ranked = rank_candidates(
            roster=roster,
            patterns=patterns,
            weekday=open_shift.date.weekday(),
            department=open_shift.department,
            signature=(
                open_shift.source_signature
                or open_shift.display_time.replace("–", "-")
            ),
            current_hours=current_hours,
            current_days=current_days,
            shift_date=open_shift.date,
        )

        top_choices = [
            {
                "employee_id": item["pattern"].employee_id,
                "name": item["pattern"].employee.full_name,
                "reasons": item["reasons"],
                "possible_split": item["possible_split"],
            }
            for item in ranked[:5]
        ]

        other_available = [
            item["pattern"].employee
            for item in ranked[5:]
        ]

        open_choice_groups.setdefault(
            open_shift.department,
            [],
        ).append(
            {
                "shift": open_shift,
                "choices": top_choices,
                "other_available": other_available,
            }
        )

    open_shift_groups = [
        {
            "department": Department.RESTAURANT,
            "label": "Restaurant",
            "items": open_choice_groups.get(
                Department.RESTAURANT,
                [],
            ),
        },
        {
            "department": Department.BAR,
            "label": "Bar",
            "items": open_choice_groups.get(
                Department.BAR,
                [],
            ),
        },
    ]
    unresolved = request.session.get(f"unresolved_{roster.pk}", [])
    unresolved_map = {
        (int(item["employee_id"]), item["date"]): item
        for item in unresolved
    }

    rows = []
    for employee in employees:
        cells = []
        for day in days:
            cells.append(
                {
                    "day": day,
                    "shifts": shift_map.get((employee.id, day), []),
                    "issue": unresolved_map.get(
                        (employee.id, day.isoformat())
                    ),
                }
            )

        pattern = patterns_by_employee.get(employee.id)
        worked_hours = round(
            employee_hours.get(employee.id, 0),
            2,
        )

        if pattern:
            if employee.department == Department.BAR:
                learned_target = round(float(pattern.bar_target_hours), 1)
            else:
                learned_target = round(float(pattern.restaurant_target_hours), 1)
            payroll_target = round(float(pattern.payroll_average_hours), 1)
        else:
            learned_target = None
            payroll_target = None

        if learned_target is None or learned_target < 8:
            allocation_status = ""
        elif worked_hours < learned_target * 0.80:
            allocation_status = "Needs hours"
        elif worked_hours > learned_target + 2.5:
            allocation_status = "Over target"
        else:
            allocation_status = "On target"

        rows.append(
            {
                "employee": employee,
                "cells": cells,
                "hours": worked_hours,
                "target_hours": learned_target,
                "payroll_target_hours": payroll_target,
                "allocation_status": allocation_status,
            }
        )

    return render(
        request,
        "roster/detail.html",
        {
            "roster": roster,
            "days": days,
            "rows": rows,
            "departments": Department.choices,
            "unresolved_count": len(unresolved),
            "open_shift_groups": open_shift_groups,
            "show_all": show_all,
            "scheduled_employee_count": len(scheduled_employee_ids),
        },
    )




@login_required
def roster_excel(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)
    days = [
        roster.week_start + timedelta(days=index)
        for index in range(7)
    ]
    shifts = list(
        roster.shifts.select_related("employee").order_by(
            "department",
            "employee__first_name",
            "employee__last_name",
            "date",
            "segment",
        )
    )
    patterns = list(
        EmployeePattern.objects.select_related("employee")
    )

    # Existing roster, keyed by employee/day.
    shift_map = {}
    current_hours = {}
    current_days = {}
    employee_days = set()
    scheduled_ids = set()
    department_hours = {}

    for shift in shifts:
        shift_map.setdefault(
            (shift.employee_id, shift.date),
            [],
        ).append(shift)
        scheduled_ids.add(shift.employee_id)
        key = (shift.employee_id, shift.department)
        department_hours[key] = (
            department_hours.get(key, 0.0)
            + shift.duration_hours
        )
        current_hours[key] = (
            current_hours.get(key, 0.0)
            + shift.duration_hours
        )
        employee_days.add(
            (shift.employee_id, shift.department, shift.date)
        )

    for employee_id, department, shift_date in employee_days:
        key = (employee_id, department)
        current_days[key] = current_days.get(key, 0) + 1

    # Provisional assignments live only in the workbook until import.
    provisional = {}
    provisional_days = set()
    impossible = []

    for open_shift in roster.open_shifts.all().order_by(
        "department",
        "date",
        "start_time",
    ):
        signature = (
            open_shift.source_signature
            or open_shift.display_time.replace("–", "-")
        )
        ranked = rank_candidates(
            roster=roster,
            patterns=patterns,
            weekday=open_shift.date.weekday(),
            department=open_shift.department,
            signature=signature,
            current_hours=current_hours,
            current_days=current_days,
            shift_date=open_shift.date,
        )

        chosen = None
        for item in ranked:
            employee_id = item["pattern"].employee_id
            # Keep the draft simple: do not provisionally give the same
            # employee two separate unresolved jobs on the same day.
            if (employee_id, open_shift.date) in provisional_days:
                continue
            chosen = item
            break

        if chosen is None:
            impossible.append((open_shift, signature))
            continue

        pattern = chosen["pattern"]
        scheduled_ids.add(pattern.employee_id)
        provisional.setdefault(
            (pattern.employee_id, open_shift.date),
            [],
        ).append(signature)
        provisional_days.add(
            (pattern.employee_id, open_shift.date)
        )

        key = (pattern.employee_id, open_shift.department)
        draft_hours = signature_duration(signature)
        department_hours[key] = (
            department_hours.get(key, 0.0)
            + draft_hours
        )
        current_hours[key] = (
            current_hours.get(key, 0.0)
            + draft_hours
        )
        current_days[key] = current_days.get(key, 0) + 1

    # Show every active employee who is actually part of the generated draft.
    employees = list(
        Employee.objects.filter(
            is_active=True,
            id__in=scheduled_ids,
        ).order_by(
            "department",
            "first_name",
            "last_name",
        )
    )
    row_department = {}
    for employee in employees:
        restaurant_hours = department_hours.get(
            (employee.id, Department.RESTAURANT), 0.0
        )
        bar_hours = department_hours.get(
            (employee.id, Department.BAR), 0.0
        )
        if bar_hours > restaurant_hours:
            row_department[employee.id] = Department.BAR
        elif restaurant_hours > 0:
            row_department[employee.id] = Department.RESTAURANT
        else:
            row_department[employee.id] = employee.department

    output = BytesIO()
    workbook = xlsxwriter.Workbook(
        output,
        {"in_memory": True},
    )

    title_format = workbook.add_format(
        {"bold": True, "font_size": 16}
    )
    section_format = workbook.add_format(
        {"bold": True, "bg_color": "#EAF0FF", "border": 1}
    )
    header_format = workbook.add_format(
        {"bold": True, "border": 1, "align": "center"}
    )
    name_format = workbook.add_format({"bold": True, "border": 1})
    cell_format = workbook.add_format(
        {"border": 1, "align": "center", "valign": "vcenter"}
    )
    replace_format = workbook.add_format({"border": 1})
    note_format = workbook.add_format(
        {"font_color": "#667085", "italic": True}
    )

    sheet = workbook.add_worksheet("Roster")
    sheet.hide_gridlines(2)
    sheet.freeze_panes(3, 2)
    sheet.set_column("A:A", 23)
    sheet.set_column("B:B", 23)
    sheet.set_column("C:I", 15)
    sheet.set_column("J:J", 11)

    sheet.merge_range(
        "A1:J1",
        f"Week ending {roster.week_end:%d %B %Y}",
        title_format,
    )
    sheet.write(
        "A2",
        "Draft roster. Leave it as it is, or use Replace with / edit a shift and upload it back.",
        note_format,
    )

    headers = ["Employee", "Replace with"]
    headers += [f"{day:%a}\n{day:%d %b}" for day in days]
    headers.append("Hours")
    for col, header in enumerate(headers):
        sheet.write(2, col, header, header_format)

    lists_sheet = workbook.add_worksheet("_Lists")
    lists_sheet.hide()
    meta_sheet = workbook.add_worksheet("_Meta")
    meta_sheet.hide()
    meta_sheet.write("A1", "roster_id")
    meta_sheet.write("B1", roster.pk)
    meta_sheet.write("A2", "week_start")
    meta_sheet.write("B2", roster.week_start.isoformat())
    meta_sheet.write("A3", "format")
    meta_sheet.write("B3", "manager-workbook-v1")
    meta_sheet.write_row(4, 0, ["excel_row", "employee_id", "department"])

    # Replacement lists are department-specific and deliberately simple.
    replacement_ranges = {}
    list_col = 0
    for department in (Department.RESTAURANT, Department.BAR):
        if department == Department.BAR:
            candidates = Employee.objects.filter(
                is_active=True,
                can_work_bar=True,
            ).order_by("first_name", "last_name")
        else:
            candidates = Employee.objects.filter(
                is_active=True,
                can_work_restaurant=True,
            ).order_by("first_name", "last_name")

        names = [employee.full_name for employee in candidates]
        if not names:
            continue
        for list_row, name in enumerate(names):
            lists_sheet.write(list_row, list_col, name)
        col_letter = xlsxwriter.utility.xl_col_to_name(list_col)
        range_name = f"Replacement_{department.title()}"
        workbook.define_name(
            range_name,
            f"='_Lists'!${col_letter}$1:${col_letter}${len(names)}",
        )
        replacement_ranges[department] = range_name
        list_col += 1

    row = 3
    meta_row = 5
    for department, label in (
        (Department.RESTAURANT, "Restaurant"),
        (Department.BAR, "Bar"),
    ):
        department_employees = [
            employee for employee in employees
            if row_department.get(employee.id) == department
        ]
        if not department_employees:
            continue

        sheet.merge_range(row, 0, row, 9, label, section_format)
        row += 1

        for employee in department_employees:
            sheet.write(row, 0, employee.full_name, name_format)
            sheet.write_blank(row, 1, None, replace_format)
            range_name = replacement_ranges.get(department)
            if range_name:
                sheet.data_validation(
                    row, 1, row, 1,
                    {
                        "validate": "list",
                        "source": f"={range_name}",
                        "input_title": "Replace employee",
                        "input_message": "Leave blank, or choose another suitable employee.",
                    },
                )

            total_hours = 0.0
            for day_col, day in enumerate(days, start=2):
                values = []
                employee_shifts = sorted(
                    shift_map.get((employee.id, day), []),
                    key=lambda item: item.segment,
                )
                for shift in employee_shifts:
                    values.append(
                        f"{shift.start_time:%H:%M}-{shift.end_time:%H:%M}"
                    )
                    total_hours += shift.duration_hours

                for signature in provisional.get((employee.id, day), []):
                    values.append(signature)
                    total_hours += signature_duration(signature)

                sheet.write(
                    row,
                    day_col,
                    ", ".join(values) if values else "OFF",
                    cell_format,
                )

            sheet.write_number(row, 9, round(total_hours, 1), cell_format)
            meta_sheet.write_row(
                meta_row,
                0,
                [row + 1, employee.id, department],
            )
            meta_row += 1
            row += 1

    if impossible:
        # This should be rare; keep it on the same sheet without creating a
        # separate manager workflow.
        row += 1
        sheet.write(row, 0, "Unassigned", section_format)
        row += 1
        for open_shift, signature in impossible:
            sheet.write(row, 0, "UNASSIGNED", name_format)
            day_col = (open_shift.date - roster.week_start).days + 2
            if 2 <= day_col <= 8:
                sheet.write(row, day_col, signature, cell_format)
            row += 1

    workbook.close()
    output.seek(0)

    filename = f"roster-week-ending-{roster.week_end:%Y-%m-%d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@transaction.atomic
def roster_excel_import(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")
    if roster.status == RosterStatus.PUBLISHED:
        messages.error(request, "Published rosters cannot be replaced from Excel.")
        return redirect("roster:detail", pk=pk)

    upload = request.FILES.get("workbook")
    if not upload:
        messages.error(request, "Choose the completed Excel roster first.")
        return redirect("roster:detail", pk=pk)

    try:
        workbook = load_workbook(upload, data_only=False)
    except Exception:
        messages.error(request, "That file is not a readable Excel workbook.")
        return redirect("roster:detail", pk=pk)

    if "Roster" not in workbook.sheetnames or "_Meta" not in workbook.sheetnames:
        messages.error(request, "Use the Manager Workbook downloaded from this roster.")
        return redirect("roster:detail", pk=pk)

    sheet = workbook["Roster"]
    meta = workbook["_Meta"]
    if meta["B1"].value != roster.pk:
        messages.error(request, "That workbook belongs to a different roster week.")
        return redirect("roster:detail", pk=pk)

    rows = []
    meta_row = 6
    while meta.cell(meta_row, 1).value:
        excel_row = int(meta.cell(meta_row, 1).value)
        employee_id = int(meta.cell(meta_row, 2).value)
        department = str(meta.cell(meta_row, 3).value)
        original = Employee.objects.filter(pk=employee_id, is_active=True).first()
        if original is None:
            messages.error(request, "An employee in this workbook no longer exists.")
            return redirect("roster:detail", pk=pk)

        replacement_name = str(sheet.cell(excel_row, 2).value or "").strip()
        employee = original
        if replacement_name:
            candidates = Employee.objects.filter(is_active=True)
            if department == Department.BAR:
                candidates = candidates.filter(can_work_bar=True)
            else:
                candidates = candidates.filter(can_work_restaurant=True)
            matches = [
                candidate for candidate in candidates
                if candidate.full_name == replacement_name
            ]
            if len(matches) != 1:
                messages.error(
                    request,
                    f"Could not uniquely match replacement employee '{replacement_name}'.",
                )
                return redirect("roster:detail", pk=pk)
            employee = matches[0]

        day_values = []
        for day_index in range(7):
            raw = sheet.cell(excel_row, day_index + 3).value
            value = str(raw or "OFF").strip()
            day_values.append(value)

        rows.append((employee, department, day_values))
        meta_row += 1

    # Replacements may point at an employee who already has a row. That is
    # intentional: their schedules are merged on import, then checked for
    # overlaps before anything is written to the database.
    parsed_rows = []
    try:
        for employee, department, day_values in rows:
            parsed_days = []
            for day_index, value in enumerate(day_values):
                if value.lower() in {"off", "-", "none", ""}:
                    parsed_days.append([])
                    continue
                parsed_days.append(parse_signature(value))
            parsed_rows.append((employee, department, parsed_days))
    except Exception:
        messages.error(
            request,
            "One of the shift cells is not valid. Use 09:00-17:00, a comma-separated split shift, or OFF.",
        )
        return redirect("roster:detail", pk=pk)

    merged = {}
    for employee, department, parsed_days in parsed_rows:
        for day_index, parsed in enumerate(parsed_days):
            if not parsed:
                continue
            shift_date = roster.week_start + timedelta(days=day_index)
            merged.setdefault(
                (employee.id, department, shift_date),
                {"employee": employee, "segments": []},
            )["segments"].extend(parsed)

    # Validate combined rows before replacing the live draft.
    for item in merged.values():
        intervals = []
        for _segment, start, end in item["segments"]:
            start_minutes = start.hour * 60 + start.minute
            end_minutes = end.hour * 60 + end.minute
            if end_minutes <= start_minutes:
                end_minutes += 24 * 60
            intervals.append((start_minutes, end_minutes))
        intervals.sort()
        for previous, current in zip(intervals, intervals[1:]):
            if previous[1] > current[0]:
                messages.error(
                    request,
                    "A replacement would give an employee overlapping shifts. Change the replacement and upload again.",
                )
                return redirect("roster:detail", pk=pk)

    roster.shifts.all().delete()
    roster.open_shifts.all().delete()

    created = 0
    for (_employee_id, department, shift_date), item in merged.items():
        employee = item["employee"]
        for segment_index, (_segment, start, end) in enumerate(
            item["segments"], start=1
        ):
            Shift.objects.create(
                roster_week=roster,
                employee=employee,
                department=department,
                date=shift_date,
                start_time=start,
                end_time=end,
                segment=segment_index,
                source="manual",
                confidence=100,
                notes="Imported from Manager Workbook",
            )
            created += 1

    messages.success(
        request,
        f"Manager Workbook applied. {created} shift segments imported.",
    )
    return redirect("roster:detail", pk=pk)
@login_required
def save_cell(request, pk):
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")
    roster = get_object_or_404(RosterWeek, pk=pk)
    employee = get_object_or_404(Employee, pk=request.POST["employee_id"])
    date = request.POST["date"]
    text = request.POST.get("shift_text","").strip()
    Shift.objects.filter(roster_week=roster, employee=employee, date=date).delete()
    if text and text.lower() not in {"off","-","none"}:
        try:
            for segment, start, end in parse_signature(text):
                Shift.objects.create(
                    roster_week=roster, employee=employee,
                    department=request.POST.get("department") or employee.department,
                    date=date, start_time=start, end_time=end, segment=segment,
                    source="manual", confidence=100,
                )
        except Exception:
            messages.error(request, "Choose: enter 09:00-17:00, enter a split shift, or type OFF.")
            return redirect("roster:detail", pk=pk)
    key = f"unresolved_{roster.pk}"
    request.session[key] = [
        i for i in request.session.get(key, [])
        if not (int(i["employee_id"]) == employee.id and i["date"] == date)
    ]
    return redirect("roster:detail", pk=pk)

@login_required
def use_suggestion(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)
    employee = get_object_or_404(Employee, pk=request.POST["employee_id"])
    date = request.POST["date"]
    suggestion = request.POST.get("suggestion","OFF")
    for segment, start, end in parse_signature(suggestion):
        Shift.objects.create(
            roster_week=roster, employee=employee, department=employee.department,
            date=date, start_time=start, end_time=end, segment=segment,
            source="learned", confidence=int(request.POST.get("confidence",0)),
        )
    key = f"unresolved_{roster.pk}"
    request.session[key] = [
        i for i in request.session.get(key, [])
        if not (int(i["employee_id"]) == employee.id and i["date"] == date)
    ]
    return redirect("roster:detail", pk=pk)

@login_required
def roster_publish(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)
    publish_roster(roster, request.user)
    return redirect("roster:detail", pk=pk)



def _employee_can_take_open_shift(roster, open_shift, employee):
    if open_shift.department == Department.BAR:
        if not employee.can_work_bar:
            return False, "This employee cannot work Bar."
    elif not employee.can_work_restaurant:
        return False, "This employee cannot work Restaurant."

    availability = candidate_availability(
        roster=roster,
        employee=employee,
        shift_date=open_shift.date,
        signature=(
            open_shift.source_signature
            or open_shift.display_time.replace("–", "-")
        ),
    )

    if not availability["available"]:
        return False, availability["reason"]

    return True, availability["reason"]


@login_required
def assign_open_shift(request, pk, open_shift_id):
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")
    roster = get_object_or_404(RosterWeek, pk=pk)
    open_shift = get_object_or_404(OpenShift, pk=open_shift_id, roster_week=roster)
    employee = get_object_or_404(Employee, pk=request.POST["employee_id"])

    allowed, reason = _employee_can_take_open_shift(
        roster,
        open_shift,
        employee,
    )
    if not allowed:
        messages.warning(
            request,
            f"{employee.full_name} is not available: {reason}",
        )
        return redirect("roster:detail", pk=pk)


    parts = parse_signature(open_shift.source_signature or open_shift.display_time.replace("–", "-"))
    for segment, start, end in parts:
        Shift.objects.create(
            roster_week=roster,
            employee=employee,
            department=open_shift.department,
            date=open_shift.date,
            start_time=start,
            end_time=end,
            segment=segment,
            source="manager_choice",
            confidence=100,
        )
    open_shift.delete()
    messages.success(request, f"Assigned to {employee.full_name}.")
    return redirect("roster:detail", pk=pk)


@login_required
def assign_suggested_employee(request, pk, open_shift_id, employee_id):
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    roster = get_object_or_404(RosterWeek, pk=pk)
    open_shift = get_object_or_404(
        OpenShift,
        pk=open_shift_id,
        roster_week=roster,
    )
    employee = get_object_or_404(Employee, pk=employee_id)

    allowed, reason = _employee_can_take_open_shift(
        roster,
        open_shift,
        employee,
    )
    if not allowed:
        messages.warning(
            request,
            f"{employee.full_name} is not available: {reason}",
        )
        return redirect("roster:detail", pk=pk)


    parts = parse_signature(
        open_shift.source_signature
        or open_shift.display_time.replace("–", "-")
    )

    for segment, start, end in parts:
        Shift.objects.create(
            roster_week=roster,
            employee=employee,
            department=open_shift.department,
            date=open_shift.date,
            start_time=start,
            end_time=end,
            segment=segment,
            source="manager_choice",
            confidence=100,
        )

    open_shift.delete()
    messages.success(request, f"Assigned to {employee.full_name}.")
    return redirect("roster:detail", pk=pk)


@login_required
def roster_delete(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)
    shift_count = roster.shifts.count()
    open_shift_count = roster.open_shifts.count()

    if request.method == "POST":
        if roster.status == RosterStatus.PUBLISHED:
            messages.error(
                request,
                "Published rosters cannot be deleted here.",
            )
            return redirect("roster:detail", pk=roster.pk)

        purpose = roster.get_purpose_display()
        label = str(roster)
        roster.delete()

        messages.success(
            request,
            f"{label} deleted. Removed {shift_count} shifts "
            f"and {open_shift_count} open shifts.",
        )

        if purpose == "Historic roster":
            messages.info(
                request,
                "Run Learning again because the historic evidence changed.",
            )

        return redirect("roster:list")

    return render(
        request,
        "roster/delete.html",
        {
            "roster": roster,
            "shift_count": shift_count,
            "open_shift_count": open_shift_count,
        },
    )


@login_required
def roster_regenerate(request, pk):
    roster = get_object_or_404(RosterWeek, pk=pk)

    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    if roster.status == RosterStatus.PUBLISHED:
        messages.error(
            request,
            "Published rosters cannot be regenerated.",
        )
        return redirect("roster:detail", pk=roster.pk)

    if roster.purpose != RosterPurpose.WEEKLY:
        messages.error(
            request,
            "Historic and base rosters are evidence. Choose: delete it, or leave it unchanged.",
        )
        return redirect("roster:detail", pk=roster.pk)

    roster.shifts.all().delete()
    roster.open_shifts.all().delete()
    request.session.pop(f"unresolved_{roster.pk}", None)

    result = generate_business_roster(
        roster,
        uncertain_threshold=75,
    )

    messages.success(
        request,
        f"Draft regenerated. {result['created']} assigned shift segments; "
        f"{result['open']} shifts need a choice.",
    )
    return redirect("roster:detail", pk=roster.pk)
