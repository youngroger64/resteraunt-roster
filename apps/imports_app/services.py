from datetime import date, datetime, timedelta
import re
from openpyxl import load_workbook
from django.db import transaction
from django.db.models.functions import Lower
from apps.employees.models import Department, Employee
from apps.roster.models import RosterPurpose, RosterWeek, Shift

OFF_WORDS = {"", "off", "0ff", "-", "none", "holiday", "hol", "leave"}

def _text(value):
    return str(value).strip() if value is not None else ""

def _normalise_name(value):
    return " ".join(_text(value).split()).strip()

def _find_sheet(workbook):
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 15), values_only=True):
            values = [_text(v).lower().strip() for v in row]
            if any(v.startswith("mon") for v in values) and any(v.startswith("sun") for v in values):
                return sheet
    return workbook.active

def _parse_week_start(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True):
        for value in row:
            match = re.search(
                r"(?:w\s*/?\s*e|week\s+ending)\s*(\d{1,2})[\/.-](\d{1,2})[\/.-](\d{2,4})",
                _text(value), re.I,
            )
            if match:
                day, month, year = map(int, match.groups())
                if year < 100:
                    year += 2000
                return date(year, month, day) - timedelta(days=6)
    return None

def _parse_clock(raw, suffix=""):
    raw = raw.lower().strip().replace(".", ":")
    raw = re.sub(r"[^0-9:]", "", raw)
    if ":" not in raw:
        raw += ":00"
    hour, minute = map(int, raw.split(":", 1))
    suffix = suffix.lower()
    if suffix == "pm" and hour < 12:
        hour += 12
    elif suffix == "am" and hour == 12:
        hour = 0
    elif suffix in {"mn", "midnight"}:
        hour, minute = 0, 0
    if hour > 23 or minute > 59:
        raise ValueError("Invalid time")
    return datetime.strptime(f"{hour:02d}:{minute:02d}", "%H:%M").time()

def _normalise_chunk(value):
    text = value.lower().strip()
    text = text.replace("—", "-").replace("–", "-").replace(" to ", "-")
    text = text.replace("midnight", "mn")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"(?<=\d),(?=\d)", ".", text)

    # Common spreadsheet mistakes.
    text = re.sub(r"^(\d{1,2}[.:]\d{1,2})\.-", r"\1-", text)
    text = re.sub(
        r"^(\d{1,2}(?:[.:]\d{1,2})?)\-(\d{1,2})\-(\d{2})(am|pm|mn)?$",
        r"\1-\2.\3\4",
        text,
    )
    text = re.sub(
        r"^(\d{1,2}[.:]\d{1,2})[.:](\d{1,2}[.:]\d{1,2})(am|pm|mn)?$",
        r"\1-\2\3",
        text,
    )
    text = re.sub(
        r"^(\d{1,2})[.:](\d{1,2})(am|pm|mn)$",
        r"\1-\2\3",
        text,
    )
    text = re.sub(r"(\d)\.(am|pm|mn)$", r"\1\2", text)
    return text

def parse_shift_cell(value):
    original = _text(value)
    text = original.lower().strip()
    if text in OFF_WORDS:
        return [], None

    text = text.replace("&", ";")
    chunks = [part.strip() for part in re.split(r";|,\s+", text) if part.strip()]
    parsed = []

    for raw_chunk in chunks:
        chunk = _normalise_chunk(raw_chunk)

        close_match = re.match(
            r"^(\d{1,2}(?:[.:]\d{1,2})?)(am|pm)?-close$", chunk
        )
        if close_match:
            start = _parse_clock(close_match.group(1), close_match.group(2) or "")
            if not close_match.group(2) and start.hour <= 7:
                start = start.replace(hour=start.hour + 12)
            parsed.append((start, _parse_clock("1", "am")))
            continue

        match = re.match(
            r"^(\d{1,2}(?:[.:]\d{1,2})?)(am|pm)?-"
            r"(\d{1,2}(?:[.:]\d{1,2})?)(am|pm|mn)?$",
            chunk,
        )
        if not match:
            return [], f"Could not understand '{original}'"

        start_raw, start_suffix, end_raw, end_suffix = match.groups()
        try:
            start = _parse_clock(start_raw, start_suffix or "")
            end = _parse_clock(end_raw, end_suffix or "")
        except ValueError:
            return [], f"Could not understand '{original}'"

        if not start_suffix and start.hour <= 7:
            start = start.replace(hour=start.hour + 12)
        if not end_suffix and end.hour <= 7:
            end = end.replace(hour=end.hour + 12)

        parsed.append((start, end))

    return parsed, None

def _employee_for_name(name, department):
    normalised = _normalise_name(name)
    pieces = normalised.split(maxsplit=1)
    first = pieces[0]
    last = pieces[1] if len(pieces) > 1 else ""

    employee = Employee.objects.annotate(
        first_lower=Lower("first_name"),
        last_lower=Lower("last_name"),
    ).filter(
        first_lower=first.lower(),
        last_lower=last.lower(),
    ).first()

    if not employee and not last:
        employee = Employee.objects.filter(first_name__iexact=first).first()

    if not employee:
        employee = Employee.objects.create(
            first_name=first,
            last_name=last,
            department=department,
            can_work_bar=department == Department.BAR,
            can_work_restaurant=department == Department.RESTAURANT,
        )
    else:
        changed = []
        if department == Department.BAR and not employee.can_work_bar:
            employee.can_work_bar = True
            changed.append("can_work_bar")
        if department == Department.RESTAURANT and not employee.can_work_restaurant:
            employee.can_work_restaurant = True
            changed.append("can_work_restaurant")
        if changed:
            changed.append("updated_at")
            employee.save(update_fields=changed)

    return employee

@transaction.atomic
def import_employees(file_obj):
    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0, ["Workbook is empty."]

    headers = [_text(v).lower().replace(" ", "_") for v in rows[0]]
    aliases = {
        "id": ["id", "employee_id", "staff_id"],
        "first_name": ["first_name", "firstname", "forename", "first"],
        "last_name": ["last_name", "lastname", "surname"],
        "name": ["name", "employee", "employee_name"],
        "department": ["department", "area", "section"],
    }

    def index_for(key):
        for alias in aliases[key]:
            if alias in headers:
                return headers.index(alias)
        return None

    indexes = {key: index_for(key) for key in aliases}
    count = 0

    for row in rows[1:]:
        full_name = _text(row[indexes["name"]]) if indexes["name"] is not None else ""
        first = _text(row[indexes["first_name"]]) if indexes["first_name"] is not None else ""
        last = _text(row[indexes["last_name"]]) if indexes["last_name"] is not None else ""

        if full_name and not first:
            parts = full_name.split(maxsplit=1)
            first = parts[0]
            last = parts[1] if len(parts) > 1 else ""
        if not first:
            continue

        external_id = _text(row[indexes["id"]]) if indexes["id"] is not None else ""
        department_text = (
            _text(row[indexes["department"]]).lower()
            if indexes["department"] is not None
            else ""
        )
        department = Department.BAR if "bar" in department_text else Department.RESTAURANT
        defaults = {
            "first_name": first,
            "last_name": last,
            "department": department,
            "can_work_bar": department == Department.BAR,
            "can_work_restaurant": department == Department.RESTAURANT,
            "is_active": True,
        }

        if external_id:
            Employee.objects.update_or_create(external_id=external_id, defaults=defaults)
        else:
            Employee.objects.update_or_create(
                first_name=first, last_name=last, defaults=defaults
            )
        count += 1

    return count, []

@transaction.atomic
def import_roster(file_obj, week_start=None, purpose=RosterPurpose.HISTORIC):
    workbook = load_workbook(file_obj, data_only=True)
    sheet = _find_sheet(workbook)
    week_start = week_start or _parse_week_start(sheet)
    if week_start is None:
        raise ValueError("No readable week date.")

    roster, _ = RosterWeek.objects.get_or_create(
        week_start=week_start,
        defaults={"purpose": purpose},
    )
    roster.purpose = purpose
    roster.save(update_fields=["purpose", "updated_at"])
    roster.shifts.all().delete()

    rows = list(sheet.iter_rows(values_only=True))
    header_index = None
    name_column = None
    day_columns = {}

    for row_index, row in enumerate(rows):
        values = [_text(value).lower().strip() for value in row]
        monday = next((i for i, value in enumerate(values) if value.startswith("mon")), None)
        sunday = next((i for i, value in enumerate(values) if value.startswith("sun")), None)
        if monday is not None and sunday is not None:
            header_index = row_index
            name_column = monday - 1
            for prefix, day_number in [
                ("mon", 0), ("tue", 1), ("wed", 2), ("thu", 3),
                ("fri", 4), ("sat", 5), ("sun", 6),
            ]:
                day_columns[day_number] = next(
                    i for i, value in enumerate(values) if value.startswith(prefix)
                )
            break

    if header_index is None:
        return roster, 0, [{
            "message": "Could not find the employee and weekday columns.",
            "choices": ["Choose another file", "Enter shifts manually"],
        }]

    department = Department.RESTAURANT
    count = 0
    issues = []

    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        name = _normalise_name(row[name_column] if name_column < len(row) else "")
        if name.lower() in {"restaurant", "bar"}:
            department = Department.BAR if name.lower() == "bar" else Department.RESTAURANT
            continue
        if not name:
            continue

        employee = _employee_for_name(name, department)

        for day_number, column in day_columns.items():
            raw_value = row[column] if column < len(row) else None
            parsed_shifts, error = parse_shift_cell(raw_value)
            if error:
                issues.append({
                    "row": row_number,
                    "employee": employee.full_name,
                    "date": (week_start + timedelta(days=day_number)).isoformat(),
                    "value": _text(raw_value),
                    "message": error,
                    "choices": ["Treat as OFF", "Fix manually"],
                })
                continue

            for segment, (start, end) in enumerate(parsed_shifts, start=1):
                Shift.objects.create(
                    roster_week=roster,
                    employee=employee,
                    department=department,
                    date=week_start + timedelta(days=day_number),
                    segment=segment,
                    start_time=start,
                    end_time=end,
                    source="imported",
                    confidence=95,
                )
                count += 1

    return roster, count, issues
